from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pyautogui
import re
from time import sleep
from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
import customtkinter as ctk
from tkinter import messagebox

def obter_valor_dolar():
    driver = webdriver.Chrome()
    driver.get("https://wise.com/br/currency-converter/dolar-hoje")

    # Espera até o campo de valor do dólar ser visível e acessível
    valor_dolar_element = WebDriverWait(driver, 10).until(
        EC.visibility_of_element_located((By.CSS_SELECTOR, "#target-input"))
    )
    
    # Extrai o valor do dólar
    valor_dolar = driver.find_element(By.CSS_SELECTOR, "#target-input").get_attribute("value")
    valor_dolar = valor_dolar.replace(',', '.')
    driver.quit()

    return float(valor_dolar)

def raspar_best_buy(produto, valor_dolar):

    print(f"Valor do dólar: {valor_dolar}")

    # Configura o navegador
    driver = webdriver.Chrome()
    driver.get('https://www.bestbuy.com/?intl=nosplash')

    # Pesquisa pelo produto fornecido
    pesquisar = WebDriverWait(driver, 5).until(
        EC.element_to_be_clickable((By.XPATH, "//input[@aria-label='Type to search. Navigate forward to hear suggestions']"))
    )
    pesquisar.click()
    pesquisar.send_keys(produto)

    pyautogui.press('enter')
    sleep(5)

    # Verifica se o arquivo Excel já existe
    try:
        wb = load_workbook('produtos_macbook_bestbuy.xlsx')
        ws = wb.active
    except FileNotFoundError:
        wb = Workbook()
        ws = wb.active
        ws.title = "Produtos"
        ws.append(["Nome", "Preço em Dólar", "Preço em Real"])

    while True:
        # Encontrar todos os containers de produtos (li com classe sku-item)
        items = driver.find_elements(By.CSS_SELECTOR, "li.sku-item")

        # Iterar sobre os itens encontrados
        for item in items:
            # Capturar o nome do produto
            try:
                nome = item.find_element(By.CSS_SELECTOR, "h4.sku-title a").text
            except:
                nome = 'Nome não encontrado'

            # Capturar o preço do produto
            preco_em_dolar = 'Preço não encontrado'
            preco_per_month = False  # Variável para verificar se o preço é por mês
            try:
                elements = item.find_elements(By.CSS_SELECTOR, "span[aria-hidden='true']")
                for element in elements:
                    texto = element.text
                    # Verifica se o preço contém um valor numérico
                    if re.match(r"^\$\d{1,3}(,\d{3})*(\.\d{2})?$", texto):
                        preco_em_dolar = texto
                        break
            except:
                pass

            # Verifica se há a expressão "for X months"
            try:
                price_disclaimer = item.find_element(By.CSS_SELECTOR, ".priceView-price-disclaimer__activation").text
                if "for" in price_disclaimer.lower() and "month" in price_disclaimer.lower():
                    preco_per_month = True
            except:
                pass

            # Se o preço for encontrado, processa o valor
            if preco_em_dolar != 'Preço não encontrado':
                # Remove qualquer informação adicional e converte para número
                preco_em_dolar_valor = re.sub(r"[^0-9,.]", "", preco_em_dolar)
                preco_dolar_value = float(preco_em_dolar_valor.replace('$', '').replace(',', '').strip())

                # Calcula o valor em reais
                preco_real = preco_dolar_value * valor_dolar

                if preco_per_month:
                    # Se for por mês, adiciona " per month" ao preço
                    preco_em_dolar = f"{preco_em_dolar_valor} per month"
                    preco_real = f"{preco_real:.2f} per month"  # Marca o preço em reais também como "per month"
            else:
                preco_real = 'Preço não disponível'

            # Adicionar os dados na planilha
            ws.append([nome, preco_em_dolar, preco_real])


        # Salvar os dados no arquivo Excel após cada página
        nome_arquivo = f"precos_{produto.replace('/', '_').replace('\\', '_')}.xlsx"
        wb.save(nome_arquivo)

        # Verifica se há próxima página
        try:
            next_button = driver.find_element(By.CSS_SELECTOR, "a.sku-list-page-next")
            if "aria-disabled" in next_button.get_attribute("class"):
                print("Não há mais páginas.")
                break
            else:
                next_button.click()
                sleep(5)  # Esperar a próxima página carregar
        except:
            print("Não encontrou o botão Next.")
            messagebox.showinfo("Concluido", "Raspagem concluída com sucesso!")
            break

    # Ajustar largura das colunas
    for col in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col[0].column)  # Obter a letra da coluna
        for cell in col:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        ws.column_dimensions[col_letter].width = max_length + 2

    # Salvar os dados no arquivo Excel ao final
    wb.save(nome_arquivo)

    # Fechar o navegador
    driver.quit()

def abrir_interface():
    def executar_raspagem():
        produto = entrada_produto.get()
        if produto:
            print("Buscando o valor do dólar...")
            valor_dolar = obter_valor_dolar()  # Pega o valor do dólar antes de realizar a raspagem
            print(f"Valor do dólar: {valor_dolar}")
            raspar_best_buy(produto, valor_dolar)  # Chama a raspagem com o valor do dólar
        else:
            print("Por favor, digite o nome de um produto.")

    # Configuração da interface
    ctk.set_appearance_mode("dark")
    ctk.set_default_color_theme("blue")

    app = ctk.CTk()
    app.title("Raspador Best Buy")
    app.geometry("800x600")
    app.state("zoomed")  # Abre em tela cheia

    # Widgets
    label_titulo = ctk.CTkLabel(app, text="Raspador Best Buy", font=("Arial", 24))
    label_titulo.pack(pady=20)

    label_produto = ctk.CTkLabel(app, text="Digite o produto que deseja pesquisar:")
    label_produto.pack(pady=10)

    entrada_produto = ctk.CTkEntry(app, width=400)
    entrada_produto.pack(pady=10)

    botao_raspar = ctk.CTkButton(app, text="Raspar Dados", command=executar_raspagem)
    botao_raspar.pack(pady=20)

    app.mainloop()

# Chama a interface
abrir_interface()